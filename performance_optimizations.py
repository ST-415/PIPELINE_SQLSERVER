"""
Performance Optimizations สำหรับ PIPELINE_SQLSERVER

การปรับปรุง Performance สำหรับการประมวลผลไฟล์ใหญ่:
1. การอ่านไฟล์แบบ Chunking
2. การประมวลผลแบบ Parallel Processing
3. การจัดการ Memory ที่ดีขึ้น
4. การแสดง Progress ที่ละเอียดขึ้น
5. การยกเลิกการทำงานได้
"""

import os
import gc
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Tuple, Callable, Any
import pandas as pd
import numpy as np
from pathlib import Path

class PerformanceOptimizer:
    """คลาสสำหรับการปรับปรุง Performance"""
    
    def __init__(self, log_callback: Optional[Callable] = None):
        self.log_callback = log_callback or print
        self.cancellation_token = threading.Event()
        self.chunk_size = 10000  # ขนาด chunk สำหรับการอ่านไฟล์
        self.max_workers = min(4, os.cpu_count() or 1)  # จำนวน worker threads
        
    def set_cancellation_token(self, token: threading.Event):
        """ตั้งค่า cancellation token สำหรับการยกเลิกการทำงาน"""
        self.cancellation_token = token
        
    def read_large_file_chunked(self, file_path: str, file_type: str = 'excel') -> Tuple[bool, pd.DataFrame]:
        """
        อ่านไฟล์ใหญ่แบบ chunked เพื่อประหยัด memory
        
        Args:
            file_path: ที่อยู่ไฟล์
            file_type: ประเภทไฟล์ ('excel', 'excel_xls', หรือ 'csv')
            
        Returns:
            Tuple[bool, pd.DataFrame]: (สำเร็จหรือไม่, DataFrame)
        """
        try:
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            self.log_callback(f"📂 กำลังอ่านไฟล์: {os.path.basename(file_path)} ({file_size_mb:.1f} MB)")
            
            if file_size_mb > 100:  # ไฟล์ใหญ่กว่า 100MB
                self.log_callback(f"⚠️ ไฟล์ขนาดใหญ่ - ใช้การอ่านแบบ chunked")
                return self._read_large_file_chunked(file_path, file_type)
            else:
                return self._read_small_file(file_path, file_type)
                
        except Exception as e:
            error_msg = f"❌ เกิดข้อผิดพลาดขณะอ่านไฟล์: {e}"
            self.log_callback(error_msg)
            return False, pd.DataFrame()
    
    def _read_small_file(self, file_path: str, file_type: str) -> Tuple[bool, pd.DataFrame]:
        """อ่านไฟล์ขนาดเล็กแบบปกติ"""
        try:
            if file_type == 'csv':
                df = pd.read_csv(file_path, header=0, encoding='utf-8')
            elif file_type == 'excel_xls':
                df = pd.read_excel(file_path, header=0, sheet_name=0, engine='xlrd')
            else:
                df = pd.read_excel(file_path, header=0, sheet_name=0, engine='openpyxl')
            
            self.log_callback(f"✅ อ่านไฟล์สำเร็จ - {len(df):,} แถว, {len(df.columns)} คอลัมน์")
            return True, df
            
        except Exception as e:
            self.log_callback(f"❌ เกิดข้อผิดพลาด: {e}")
            return False, pd.DataFrame()
    
    def _read_large_file_chunked(self, file_path: str, file_type: str) -> Tuple[bool, pd.DataFrame]:
        """อ่านไฟล์ใหญ่แบบ chunked"""
        try:
            chunks = []
            total_rows = 0
            
            if file_type == 'csv':
                # นับจำนวนแถวทั้งหมดก่อน
                total_rows = sum(1 for _ in open(file_path, 'r', encoding='utf-8')) - 1
                self.log_callback(f"📊 จำนวนแถวทั้งหมด: {total_rows:,}")
                
                # อ่านแบบ chunk
                chunk_reader = pd.read_csv(file_path, header=0, encoding='utf-8', chunksize=self.chunk_size)
                
                for i, chunk in enumerate(chunk_reader):
                    if self.cancellation_token.is_set():
                        self.log_callback("❌ การทำงานถูกยกเลิก")
                        return False, pd.DataFrame()
                    
                    chunks.append(chunk)
                    processed_rows = (i + 1) * self.chunk_size
                    progress = min(processed_rows / total_rows, 1.0)
                    
                    self.log_callback(f"📖 อ่าน chunk {i+1}: {len(chunk):,} แถว ({progress*100:.1f}%)")
                    
                    # ปล่อย memory ทุก 10 chunks
                    if (i + 1) % 10 == 0:
                        gc.collect()
                        
            elif file_type == 'excel_xls':  # .xls file
                # สำหรับ .xls ไฟล์ใหญ่ ใช้ xlrd
                import xlrd
                
                workbook = xlrd.open_workbook(file_path)
                worksheet = workbook.sheet_by_index(0)
                
                # อ่าน header
                headers = []
                for col_idx in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(0, col_idx)
                    headers.append(cell_value)
                
                # อ่านข้อมูลแบบ chunk
                chunk_data = []
                for row_idx in range(1, worksheet.nrows):
                    if self.cancellation_token.is_set():
                        self.log_callback("❌ การทำงานถูกยกเลิก")
                        return False, pd.DataFrame()
                    
                    row_data = []
                    for col_idx in range(worksheet.ncols):
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        row_data.append(cell_value)
                    
                    chunk_data.append(row_data)
                    
                    # สร้าง chunk ทุก chunk_size แถว
                    if len(chunk_data) >= self.chunk_size:
                        chunk_df = pd.DataFrame(chunk_data, columns=headers)
                        chunks.append(chunk_df)
                        chunk_data = []
                        
                        self.log_callback(f"📖 อ่าน chunk {len(chunks)}: {len(chunk_df):,} แถว")
                        
                        # ปล่อย memory
                        gc.collect()
                
                # เพิ่มข้อมูลที่เหลือสำหรับ .xls
                if chunk_data:
                    chunk_df = pd.DataFrame(chunk_data, columns=headers)
                    chunks.append(chunk_df)
                        
            else:  # Excel .xlsx file
                # สำหรับ Excel ไฟล์ใหญ่ ให้อ่านแบบ chunk ด้วย openpyxl
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                worksheet = workbook.active
                
                # อ่าน header
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value)
                
                # อ่านข้อมูลแบบ chunk
                chunk_data = []
                for row_idx in range(2, worksheet.max_row + 1):
                    if self.cancellation_token.is_set():
                        self.log_callback("❌ การทำงานถูกยกเลิก")
                        workbook.close()
                        return False, pd.DataFrame()
                    
                    row_data = []
                    for col_idx in range(1, len(headers) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        row_data.append(cell.value)
                    
                    chunk_data.append(row_data)
                    
                    # สร้าง chunk ทุก chunk_size แถว
                    if len(chunk_data) >= self.chunk_size:
                        chunk_df = pd.DataFrame(chunk_data, columns=headers)
                        chunks.append(chunk_df)
                        chunk_data = []
                        
                        self.log_callback(f"📖 อ่าน chunk {len(chunks)}: {len(chunk_df):,} แถว")
                        
                        # ปล่อย memory
                        gc.collect()
                
                # เพิ่มข้อมูลที่เหลือ
                if chunk_data:
                    chunk_df = pd.DataFrame(chunk_data, columns=headers)
                    chunks.append(chunk_df)
                
                if file_type != 'excel_xls':
                    workbook.close()
            
            # รวม chunks
            if chunks:
                self.log_callback("🔄 กำลังรวม chunks...")
                df = pd.concat(chunks, ignore_index=True)
                del chunks  # ปล่อย memory
                gc.collect()
                
                self.log_callback(f"✅ อ่านไฟล์สำเร็จ - {len(df):,} แถว, {len(df.columns)} คอลัมน์")
                return True, df
            else:
                self.log_callback("❌ ไม่มีข้อมูลในไฟล์")
                return False, pd.DataFrame()
                
        except Exception as e:
            self.log_callback(f"❌ เกิดข้อผิดพลาดในการอ่านไฟล์แบบ chunked: {e}")
            return False, pd.DataFrame()
    
    def process_dataframe_in_chunks(self, df: pd.DataFrame, chunk_size: int = 5000) -> List[pd.DataFrame]:
        """
        แบ่ง DataFrame เป็น chunks สำหรับการประมวลผล
        
        Args:
            df: DataFrame ที่ต้องการแบ่ง
            chunk_size: ขนาดของแต่ละ chunk
            
        Returns:
            List[pd.DataFrame]: รายการ DataFrame chunks
        """
        chunks = []
        total_rows = len(df)
        
        for i in range(0, total_rows, chunk_size):
            end_idx = min(i + chunk_size, total_rows)
            chunk = df.iloc[i:end_idx].copy()
            chunks.append(chunk)
            
            if self.cancellation_token.is_set():
                break
        
        return chunks
    
    def parallel_process_files(self, file_paths: List[str], process_func: Callable, 
                             progress_callback: Optional[Callable] = None) -> List[Tuple[bool, Any]]:
        """
        ประมวลผลไฟล์หลายไฟล์แบบ parallel
        
        Args:
            file_paths: รายการที่อยู่ไฟล์
            process_func: ฟังก์ชันสำหรับประมวลผลไฟล์
            progress_callback: callback สำหรับแสดงความคืบหน้า
            
        Returns:
            List[Tuple[bool, Any]]: ผลลัพธ์การประมวลผล
        """
        results = []
        completed = 0
        total_files = len(file_paths)
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # ส่งงานไปยัง executor
            future_to_file = {
                executor.submit(process_func, file_path): file_path 
                for file_path in file_paths
            }
            
            # รอผลลัพธ์
            for future in as_completed(future_to_file):
                if self.cancellation_token.is_set():
                    self.log_callback("❌ การทำงานถูกยกเลิก")
                    break
                
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append((True, result))
                except Exception as e:
                    self.log_callback(f"❌ เกิดข้อผิดพลาดในการประมวลผล {os.path.basename(file_path)}: {e}")
                    results.append((False, str(e)))
                
                completed += 1
                if progress_callback:
                    progress = completed / total_files
                    progress_callback(progress, f"ประมวลผลไฟล์ {completed}/{total_files}")
        
        return results
    
    def optimize_memory_usage(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        ปรับปรุงการใช้ memory ของ DataFrame
        
        Args:
            df: DataFrame ที่ต้องการปรับปรุง
            
        Returns:
            pd.DataFrame: DataFrame ที่ปรับปรุงแล้ว
        """
        try:
            initial_memory = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
            self.log_callback(f"💾 Memory เริ่มต้น: {initial_memory:.2f} MB")
            
            # ลดขนาดของ numeric columns
            for col in df.select_dtypes(include=['int64']).columns:
                df[col] = pd.to_numeric(df[col], downcast='integer')
            
            for col in df.select_dtypes(include=['float64']).columns:
                df[col] = pd.to_numeric(df[col], downcast='float')
            
            # ลดขนาดของ object columns
            for col in df.select_dtypes(include=['object']).columns:
                if df[col].nunique() / len(df) < 0.5:  # ถ้า unique values น้อยกว่า 50%
                    df[col] = df[col].astype('category')
            
            final_memory = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
            memory_saved = initial_memory - final_memory
            
            self.log_callback(f"💾 Memory หลังปรับปรุง: {final_memory:.2f} MB (ประหยัด {memory_saved:.2f} MB)")
            
            return df
            
        except Exception as e:
            self.log_callback(f"⚠️ ไม่สามารถปรับปรุง memory ได้: {e}")
            return df
    
    def create_progress_tracker(self, total_items: int, description: str = "") -> Callable:
        """
        สร้าง progress tracker สำหรับการติดตามความคืบหน้า
        
        Args:
            total_items: จำนวนรายการทั้งหมด
            description: คำอธิบายการทำงาน
            
        Returns:
            Callable: ฟังก์ชันสำหรับอัปเดตความคืบหน้า
        """
        start_time = time.time()
        completed = 0
        
        def update_progress(items_completed: int = 1, custom_message: str = ""):
            nonlocal completed
            completed += items_completed
            
            if total_items > 0:
                progress = completed / total_items
                elapsed_time = time.time() - start_time
                
                if completed > 0:
                    estimated_total = elapsed_time / completed * total_items
                    remaining_time = estimated_total - elapsed_time
                    
                    message = custom_message or f"{description}: {completed}/{total_items}"
                    time_info = f" (เหลือ {remaining_time:.1f}s)"
                    
                    return progress, message + time_info
                else:
                    return 0.0, f"{description}: เริ่มต้น..."
            else:
                return 1.0, f"{description}: เสร็จสิ้น"
        
        return update_progress
    
    def cleanup_memory(self):
        """ทำความสะอาด memory"""
        gc.collect()
        self.log_callback("🧹 ทำความสะอาด memory เรียบร้อย")


class LargeFileProcessor:
    """คลาสสำหรับการประมวลผลไฟล์ใหญ่โดยเฉพาะ"""
    
    def __init__(self, log_callback: Optional[Callable] = None):
        self.optimizer = PerformanceOptimizer(log_callback)
        self.log_callback = log_callback or print
        
    def process_large_file(self, file_path: str, file_type: str, 
                          processing_steps: List[Callable]) -> Tuple[bool, pd.DataFrame]:
        """
        ประมวลผลไฟล์ใหญ่ด้วยขั้นตอนที่กำหนด
        
        Args:
            file_path: ที่อยู่ไฟล์
            file_type: ประเภทไฟล์
            processing_steps: รายการฟังก์ชันสำหรับประมวลผล
            
        Returns:
            Tuple[bool, pd.DataFrame]: (สำเร็จหรือไม่, DataFrame ที่ประมวลผลแล้ว)
        """
        try:
            # ขั้นตอนที่ 1: อ่านไฟล์
            success, df = self.optimizer.read_large_file_chunked(file_path, file_type)
            if not success:
                return False, pd.DataFrame()
            
            # ขั้นตอนที่ 2: ปรับปรุง memory
            df = self.optimizer.optimize_memory_usage(df)
            
            # ขั้นตอนที่ 3: ประมวลผลตามขั้นตอนที่กำหนด
            for i, step_func in enumerate(processing_steps):
                if self.optimizer.cancellation_token.is_set():
                    self.log_callback("❌ การทำงานถูกยกเลิก")
                    return False, pd.DataFrame()
                
                self.log_callback(f"🔄 ขั้นตอน {i+1}/{len(processing_steps)}: {step_func.__name__}")
                df = step_func(df)
                
                # ทำความสะอาด memory หลังแต่ละขั้นตอน
                self.optimizer.cleanup_memory()
            
            return True, df
            
        except Exception as e:
            self.log_callback(f"❌ เกิดข้อผิดพลาดในการประมวลผลไฟล์: {e}")
            return False, pd.DataFrame()
    
    def set_cancellation_token(self, token: threading.Event):
        """ตั้งค่า cancellation token"""
        self.optimizer.set_cancellation_token(token)


# ฟังก์ชันช่วยเหลือสำหรับการประมวลผล
def create_chunk_processor(chunk_size: int = 5000):
    """สร้างฟังก์ชันสำหรับประมวลผลข้อมูลแบบ chunk"""
    def process_in_chunks(df: pd.DataFrame, process_func: Callable) -> pd.DataFrame:
        """ประมวลผล DataFrame แบบ chunk"""
        results = []
        total_chunks = (len(df) + chunk_size - 1) // chunk_size
        
        for i in range(0, len(df), chunk_size):
            chunk = df.iloc[i:i+chunk_size].copy()
            processed_chunk = process_func(chunk)
            results.append(processed_chunk)
            
            # แสดงความคืบหน้า
            chunk_num = (i // chunk_size) + 1
            print(f"📊 ประมวลผล chunk {chunk_num}/{total_chunks}")
        
        return pd.concat(results, ignore_index=True)
    
    return process_in_chunks


def estimate_processing_time(file_size_mb: float, processing_type: str = 'standard') -> float:
    """
    ประมาณเวลาการประมวลผล
    
    Args:
        file_size_mb: ขนาดไฟล์ใน MB
        processing_type: ประเภทการประมวลผล
        
    Returns:
        float: เวลาที่ประมาณการ (วินาที)
    """
    # อัตราการประมวลผลโดยประมาณ (MB/วินาที)
    rates = {
        'fast': 10.0,      # 10 MB/วินาที
        'standard': 5.0,   # 5 MB/วินาที
        'slow': 2.0        # 2 MB/วินาที
    }
    
    rate = rates.get(processing_type, rates['standard'])
    estimated_time = file_size_mb / rate
    
    return estimated_time


def format_file_size(size_bytes: int) -> str:
    """แปลงขนาดไฟล์เป็นรูปแบบที่อ่านง่าย"""
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"


def format_time(seconds: float) -> str:
    """แปลงเวลาเป็นรูปแบบที่อ่านง่าย"""
    if seconds < 60:
        return f"{seconds:.1f} วินาที"
    elif seconds < 3600:
        minutes = seconds / 60
        return f"{minutes:.1f} นาที"
    else:
        hours = seconds / 3600
        return f"{hours:.1f} ชั่วโมง" 